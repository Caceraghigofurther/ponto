import socket
import json
import os
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "registro_ponto.xlsx"
DB_FILE = "registro_ponto.db"

def init_excel():
    # Se o arquivo não existir, cria com cabeçalhos
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Username", "Data", "Hora", "Windows Info"])
        wb.save(EXCEL_FILE)

def init_db():
    # Inicializa o banco de dados SQLite
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ponto (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            data TEXT,
            hora TEXT,
            windows_info TEXT
        )
    """)
    conn.commit()
    conn.close()

def check_duplicate(username, date_str):
    # Verifica no arquivo Excel se já existe um registro para o usuário na data informada
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == date_str:
            return True
    return False

def register_ponto(data):
    username = data.get("username")
    windows_info = data.get("windows_info")
    timestamp = data.get("timestamp")
    dt = datetime.fromisoformat(timestamp)
    date_str = dt.strftime("%Y-%m-%d")
    time_str = dt.strftime("%H:%M:%S")
    
    if check_duplicate(username, date_str):
        return {"status": "error", "message": "Já foi registrado o ponto para hoje."}
    
    # Salva no Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([username, date_str, time_str, windows_info])
    wb.save(EXCEL_FILE)
    
    # Salva no banco de dados SQLite
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO ponto (username, data, hora, windows_info) VALUES (?, ?, ?, ?)", 
                   (username, date_str, time_str, windows_info))
    conn.commit()
    conn.close()
    
    return {"status": "success", "message": "Ponto registrado com sucesso."}

def main():
    init_excel()
    init_db()
    host = '0.0.0.0'
    port = 5000
    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_socket.bind((host, port))
    server_socket.listen(5)
    print("Servidor de ponto rodando na porta", port)
    
    while True:
        client_socket, addr = server_socket.accept()
        print("Conexão de", addr)
        data = client_socket.recv(1024).decode("utf-8")
        if not data:
            client_socket.close()
            continue
        try:
            data = json.loads(data)
            response = register_ponto(data)
        except Exception as e:
            response = {"status": "error", "message": "Erro: " + str(e)}
        client_socket.send(json.dumps(response).encode("utf-8"))
        client_socket.close()

if __name__ == "__main__":
    main()
